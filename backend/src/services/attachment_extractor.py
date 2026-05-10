"""Extract text from PDF and Excel attachments."""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def extract_text(file_path: Path, content_type: str | None) -> str | None:
    """Extract text from a file based on its content type.

    Returns extracted text or None if not extractable.
    """
    if content_type is None:
        suffix = file_path.suffix.lower()
        if suffix == ".pdf":
            content_type = "application/pdf"
        elif suffix in (".xlsx", ".xls"):
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if content_type == "application/pdf":
        return _extract_pdf(file_path)
    elif content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return _extract_excel(file_path)

    return None


def _extract_pdf(file_path: Path) -> str | None:
    try:
        import pymupdf

        doc = pymupdf.open(str(file_path))
        text_parts = []
        for page in doc:
            text_parts.append(page.get_text())
        doc.close()
        text = "\n".join(text_parts).strip()
        return text if text else None
    except Exception:
        logger.exception("Failed to extract text from PDF: %s", file_path)
        return None


def _extract_excel(file_path: Path) -> str | None:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    text_parts.append("\t".join(cells))
        wb.close()
        text = "\n".join(text_parts).strip()
        return text if text else None
    except Exception:
        logger.exception("Failed to extract text from Excel: %s", file_path)
        return None
